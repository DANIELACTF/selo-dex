"""Testa a parte 2 do onboarding: planilha de particularidades, aritmética
de competências e a carência de 3 competências antes da distribuição."""
from pathlib import Path

import openpyxl
import pytest

from onboarding.alimentar_carteira import COL_COMPETENCIA, COL_LIBERA, alimentar
from onboarding.competencia import (
    MESES_CARENCIA,
    competencia_liberacao,
    competencias_restantes,
    liberada,
    somar_meses,
    validar,
)
from onboarding.planilha_particularidades import COLUNAS, gerar

CSV_CABECALHO = (
    "numero,empresa,cnpj,regime_informado,regime_simples_api,divergencia_regime,"
    "situacao_cadastral,certificado_recebido,alerta_certificado,email_contato,"
    "grupo_obs,data_processamento,status_distribuicao\n"
)


def _csv_pendentes(tmp_path: Path) -> Path:
    p = tmp_path / "pendentes.csv"
    p.write_text(
        CSV_CABECALHO
        + "1091,C LORENA DISTRIBUIDORA LTDA,68.717.251/0001-25,Lucro Real,?,,,Sim,,a@b.com,,2026-08-28,Pendente\n"
        + "1092,GRAFICA SQUARE LTDA,30.569.577/0001-80,Simples Nacional,?,,,Não,SIM,c@d.com,,2026-08-28,Pendente\n",
        encoding="utf-8",
    )
    return p


def _carteira_fake(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carteira Completa"
    ws.append(["N° Cliente", "Nome", "CNPJ", "Regime Tributário", "Segmento",
               "Analista Responsável", "Nível", "Status"])
    ws.append(["500", "EMPRESA ANTIGA LTDA", "11.111.111/0001-11", "Simples Nacional",
               "Serviço", "Wellington", "Auxiliar", "OK"])
    ws.auto_filter.ref = "A1:H2"

    wsp = wb.create_sheet("Pendentes Daniela")
    wsp.append(["N° Cliente", "Nome", "CNPJ", "Regime Tributário", "Segmento",
                "Sugestão Analista", "Origem", "Observação"])

    wsr = wb.create_sheet("Resumo Equipe")
    wsr["A3"] = "Wellington"
    wsr["C3"] = "=COUNTIF('Carteira Completa'!$F:$F,$A3)"

    caminho = tmp_path / "carteira.xlsx"
    wb.save(caminho)
    return caminho


def _preencher(planilha: Path, linha: int, valores: dict) -> None:
    wb = openpyxl.load_workbook(planilha)
    ws = wb["Particularidades"]
    idx = {str(c.value).strip(): i for i, c in enumerate(next(ws.iter_rows(max_row=1)), 1) if c.value}
    for coluna, valor in valores.items():
        ws.cell(row=linha, column=idx[coluna], value=valor)
    wb.save(planilha)


# --------------------------------------------------------------- competência

def test_somar_meses_vira_o_ano():
    assert somar_meses("11/2026", 3) == "02/2027"
    assert somar_meses("12/2026", 1) == "01/2027"
    assert somar_meses("01/2026", 0) == "01/2026"


def test_liberacao_e_tres_competencias_apos_a_entrada():
    assert MESES_CARENCIA == 3
    assert competencia_liberacao("08/2026") == "11/2026"


def test_carencia_nao_vence_antes_da_terceira_competencia():
    assert liberada("08/2026", "08/2026") is False
    assert liberada("08/2026", "09/2026") is False
    assert liberada("08/2026", "10/2026") is False
    assert liberada("08/2026", "11/2026") is True
    assert liberada("08/2026", "12/2026") is True


def test_competencias_restantes():
    assert competencias_restantes("08/2026", "08/2026") == 3
    assert competencias_restantes("08/2026", "10/2026") == 1
    assert competencias_restantes("08/2026", "11/2026") == 0
    assert competencias_restantes("08/2026", "12/2026") == 0


@pytest.mark.parametrize("valor", ["", "13/2026", "2026-08", "agosto/2026", "8-2026"])
def test_competencia_invalida_e_recusada(valor):
    with pytest.raises(ValueError):
        validar(valor)


# ----------------------------------------------------------------- planilha

def test_planilha_tem_abas_e_prefixa_identificacao(tmp_path):
    saida = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    wb = openpyxl.load_workbook(saida)
    assert wb.sheetnames == ["Instruções", "Particularidades", "Listas"]

    ws = wb["Particularidades"]
    assert ws.max_row == 3
    assert ws.max_column == len(COLUNAS)
    assert [ws.cell(row=2, column=c).value for c in (1, 2, 3)] == [
        "1091", "C LORENA DISTRIBUIDORA LTDA", "68.717.251/0001-25"
    ]
    assert ws.freeze_panes == "D2"


def test_planilha_preenche_competencia_de_entrada(tmp_path):
    saida = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    ws = openpyxl.load_workbook(saida)["Particularidades"]
    idx = {str(c.value).strip(): i for i, c in enumerate(next(ws.iter_rows(max_row=1)), 1) if c.value}
    assert ws.cell(row=2, column=idx[COL_COMPETENCIA]).value == "08/2026"


def test_planilha_traz_certificado_e_regime_da_triagem(tmp_path):
    saida = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    ws = openpyxl.load_workbook(saida)["Particularidades"]
    idx = {str(c.value).strip(): i for i, c in enumerate(next(ws.iter_rows(max_row=1)), 1) if c.value}
    assert ws.cell(row=2, column=idx["Certificado A1"]).value == "recebido"
    assert ws.cell(row=3, column=idx["Certificado A1"]).value == "pendente"
    assert ws.cell(row=2, column=idx["Regime confirmado"]).value == "Lucro Real"


def test_planilha_exige_triagem_previa(tmp_path):
    with pytest.raises(SystemExit):
        gerar(tmp_path / "nao-existe.csv", tmp_path / "saida.xlsx")


# ------------------------------------------------------- entrada em carência

def test_empresa_nova_entra_em_carencia_e_nao_na_carteira(tmp_path):
    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    _preencher(planilha, 2, {"Responsável (analista)": "Matheus Telles", "Nível / equipe": "Sênior"})

    r = alimentar(planilha, _carteira_fake(tmp_path), tmp_path / "nova.xlsx", referencia="08/2026")

    assert r["distribuidas"] == []
    assert len(r["entraram_carencia"]) == 2
    assert "libera em 11/2026" in r["entraram_carencia"][0]

    wb = openpyxl.load_workbook(r["saida"])
    assert wb["Carteira Completa"].max_row == 2  # só a empresa antiga
    assert wb["Pendentes Daniela"].max_row == 3  # cabeçalho + 2


def test_pendentes_ganha_colunas_de_controle(tmp_path):
    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    r = alimentar(planilha, _carteira_fake(tmp_path), tmp_path / "nova.xlsx", referencia="08/2026")

    wsp = openpyxl.load_workbook(r["saida"])["Pendentes Daniela"]
    idx = {str(c.value).strip(): i for i, c in enumerate(next(wsp.iter_rows(max_row=1)), 1) if c.value}
    assert COL_COMPETENCIA in idx and COL_LIBERA in idx
    assert wsp.cell(row=2, column=idx[COL_COMPETENCIA]).value == "08/2026"
    assert wsp.cell(row=2, column=idx[COL_LIBERA]).value == "11/2026"


def test_nao_duplica_quem_ja_esta_em_carencia(tmp_path):
    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    carteira = _carteira_fake(tmp_path)

    primeira = alimentar(planilha, carteira, tmp_path / "n1.xlsx", referencia="08/2026")
    segunda = alimentar(planilha, primeira["saida"], tmp_path / "n2.xlsx", referencia="09/2026")

    assert segunda["entraram_carencia"] == []
    assert len(segunda["em_carencia"]) == 2
    assert openpyxl.load_workbook(segunda["saida"])["Pendentes Daniela"].max_row == 3


# ------------------------------------------------------------- distribuição

def test_distribui_somente_apos_tres_competencias(tmp_path):
    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    _preencher(planilha, 2, {"Responsável (analista)": "Matheus Telles", "Nível / equipe": "Sênior",
                             "Segmento": "Comércio"})
    _preencher(planilha, 3, {"Responsável (analista)": "Dulce Neves", "Nível / equipe": "Júnior"})

    entrada = alimentar(planilha, _carteira_fake(tmp_path), tmp_path / "n1.xlsx", referencia="08/2026")

    # 10/2026: ainda dentro da carência
    antes = alimentar(planilha, entrada["saida"], tmp_path / "n2.xlsx", referencia="10/2026")
    assert antes["distribuidas"] == []
    assert len(antes["em_carencia"]) == 2

    # 11/2026: carência vencida
    depois = alimentar(planilha, entrada["saida"], tmp_path / "n3.xlsx", referencia="11/2026")
    assert sorted(depois["distribuidas"]) == ["1091", "1092"]

    wb = openpyxl.load_workbook(depois["saida"])
    assert wb["Pendentes Daniela"].max_row == 1  # só o cabeçalho
    linha = [c.value for c in wb["Carteira Completa"][3]]
    assert linha[0] == "1091"
    assert linha[5] == "Matheus Telles"
    assert linha[7] == "🆕 Novo"


def test_liberada_sem_responsavel_nao_e_distribuida(tmp_path):
    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    entrada = alimentar(planilha, _carteira_fake(tmp_path), tmp_path / "n1.xlsx", referencia="08/2026")

    r = alimentar(planilha, entrada["saida"], tmp_path / "n2.xlsx", referencia="11/2026")

    assert r["distribuidas"] == []
    assert len(r["liberadas_sem_responsavel"]) == 2
    assert openpyxl.load_workbook(r["saida"])["Pendentes Daniela"].max_row == 3


def test_linha_antiga_sem_competencia_nao_e_tocada(tmp_path):
    carteira = _carteira_fake(tmp_path)
    wb = openpyxl.load_workbook(carteira)
    wb["Pendentes Daniela"].append(["999", "CLIENTE ANTIGO LTDA", "22.222.222/0001-22",
                                    "Simples Nacional", "Serviço", "Wellington", "Carteira", ""])
    wb.save(carteira)

    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    r = alimentar(planilha, carteira, tmp_path / "nova.xlsx", referencia="12/2027")

    assert "999" not in r["distribuidas"]
    nums = [str(c[0].value) for c in openpyxl.load_workbook(r["saida"])["Pendentes Daniela"].iter_rows(min_row=2)]
    assert "999" in nums


def test_competencia_invalida_na_planilha_e_reportada(tmp_path):
    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    _preencher(planilha, 2, {COL_COMPETENCIA: "13/2026"})

    r = alimentar(planilha, _carteira_fake(tmp_path), tmp_path / "nova.xlsx", referencia="08/2026")

    assert len(r["competencia_invalida"]) == 1
    assert "1091" in r["competencia_invalida"][0]
    assert not any("1091" in e for e in r["entraram_carencia"])


# ------------------------------------------------------------------ garantias

def test_empresa_ja_na_carteira_e_ignorada(tmp_path):
    carteira = _carteira_fake(tmp_path)
    wb = openpyxl.load_workbook(carteira)
    wb["Carteira Completa"].append(["1091", "C LORENA DISTRIBUIDORA LTDA", "68.717.251/0001-25",
                                    "Lucro Real", "Comércio", "Matheus Telles", "Sênior", "OK"])
    wb.save(carteira)

    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    r = alimentar(planilha, carteira, tmp_path / "nova.xlsx", referencia="08/2026")

    assert r["ja_na_carteira"] == ["1091"]
    assert not any("1091" in e for e in r["entraram_carencia"])


def test_preserva_formulas_do_resumo(tmp_path):
    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    r = alimentar(planilha, _carteira_fake(tmp_path), tmp_path / "nova.xlsx", referencia="08/2026")

    wsr = openpyxl.load_workbook(r["saida"], data_only=False)["Resumo Equipe"]
    assert wsr["C3"].value == "=COUNTIF('Carteira Completa'!$F:$F,$A3)"


def test_carteira_original_nao_e_alterada(tmp_path):
    planilha = gerar(_csv_pendentes(tmp_path), tmp_path / "part.xlsx", "08/2026")
    carteira = _carteira_fake(tmp_path)
    antes = carteira.read_bytes()

    alimentar(planilha, carteira, tmp_path / "nova.xlsx", referencia="08/2026")
    assert carteira.read_bytes() == antes
