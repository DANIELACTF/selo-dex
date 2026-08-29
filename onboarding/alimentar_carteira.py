"""Alimenta a Carteira Tributária Fiscal respeitando a carência.

Empresa nova não entra direto na carteira do analista. Ela cumpre
`MESES_CARENCIA` competências sob a Gestão Fiscal, na aba "Pendentes
Daniela", e só depois é distribuída. A rotina faz os dois movimentos:

1. **Entrada em carência** — empresa da planilha que ainda não está em
   lugar nenhum entra em "Pendentes Daniela" com a competência de entrada
   e a competência em que libera.
2. **Distribuição** — empresa em "Pendentes Daniela" cuja carência já
   venceu E que tenha responsável definido na planilha sai de lá e entra
   na "Carteira Completa".

Grava sempre em arquivo NOVO — a carteira original nunca é sobrescrita.

A aba "Resumo Equipe" é COUNTIF sobre a "Carteira Completa": recalcula
sozinha ao abrir no Excel. Empresa em carência não conta para nenhum
analista, que é o comportamento correto — ela ainda não é de ninguém.
"""
from __future__ import annotations

import copy as _copy
import datetime as dt
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from onboarding.competencia import (
    MESES_CARENCIA,
    competencia_atual,
    competencia_liberacao,
    competencias_restantes,
    liberada,
    validar,
)

ABA_CARTEIRA = "Carteira Completa"
ABA_PENDENTES = "Pendentes Daniela"
ABA_PARTICULARIDADES = "Particularidades"

STATUS_NOVO = "🆕 Novo"
ORIGEM_ONBOARDING = "🆕 Onboarding"
COL_COMPETENCIA = "Competência entrada"
COL_LIBERA = "Libera em"


def _indices(ws) -> dict[str, int]:
    return {
        str(c.value).strip(): i
        for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)
        if c.value
    }


def _garantir_colunas(ws, colunas: list[str]) -> dict[str, int]:
    """Acrescenta ao fim da aba as colunas de controle que faltarem.

    Aditivo de propósito: não mexe nas colunas que o escritório já usa.
    """
    idx = _indices(ws)
    for nome in colunas:
        if nome not in idx:
            nova = ws.max_column + 1
            cabecalho = ws.cell(row=1, column=nova, value=nome)
            modelo = ws.cell(row=1, column=1)
            cabecalho.font = _copy.copy(modelo.font)
            cabecalho.fill = _copy.copy(modelo.fill)
            ws.column_dimensions[get_column_letter(nova)].width = 18
            idx[nome] = nova
    return idx


def _ler_particularidades(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    if ABA_PARTICULARIDADES not in wb.sheetnames:
        raise SystemExit(f"A planilha {path} não tem a aba '{ABA_PARTICULARIDADES}'.")
    ws = wb[ABA_PARTICULARIDADES]
    idx = _indices(ws)
    faltando = [c for c in ("N° Cliente", "Razão social", "CNPJ") if c not in idx]
    if faltando:
        raise SystemExit(f"Colunas ausentes na planilha de particularidades: {faltando}")

    empresas = []
    for row in ws.iter_rows(min_row=2):
        def val(col: str) -> str:
            i = idx.get(col)
            v = row[i - 1].value if i else None
            return str(v).strip() if v not in (None, "") else ""

        if not val("N° Cliente"):
            continue
        empresas.append({
            "numero": val("N° Cliente"),
            "nome": val("Razão social"),
            "cnpj": val("CNPJ"),
            "regime": val("Regime confirmado") or "⚠ A confirmar",
            "segmento": val("Segmento"),
            "analista": val("Responsável (analista)"),
            "nivel": val("Nível / equipe"),
            "situacao": val("Situação"),
            "competencia": val(COL_COMPETENCIA),
            "obs": val("Obs. para a carteira"),
        })
    return empresas


def _numeros(ws, col: int) -> set[str]:
    return {
        str(r[col - 1].value).strip()
        for r in ws.iter_rows(min_row=2)
        if r[col - 1].value not in (None, "")
    }


def _estender_filtro(ws, ultima_linha: int) -> None:
    if ws.auto_filter.ref:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ultima_linha, 1)}"


def alimentar(
    particularidades: Path,
    carteira: Path,
    saida: Path,
    referencia: str | None = None,
) -> dict:
    referencia = referencia or competencia_atual()
    validar(referencia)

    empresas = _ler_particularidades(particularidades)
    por_numero = {e["numero"]: e for e in empresas}
    wb = load_workbook(carteira)  # data_only=False: preserva as fórmulas do Resumo

    for aba in (ABA_CARTEIRA, ABA_PENDENTES):
        if aba not in wb.sheetnames:
            raise SystemExit(f"A carteira {carteira} não tem a aba '{aba}'.")

    ws = wb[ABA_CARTEIRA]
    idx_cart = _indices(ws)
    wsp = wb[ABA_PENDENTES]
    idx_pend = _garantir_colunas(wsp, [COL_COMPETENCIA, COL_LIBERA])

    ja_na_carteira = _numeros(ws, idx_cart["N° Cliente"])
    ja_em_pendentes = _numeros(wsp, idx_pend["N° Cliente"])

    r: dict[str, list] = {
        "entraram_carencia": [], "distribuidas": [], "em_carencia": [],
        "liberadas_sem_responsavel": [], "ja_na_carteira": [], "competencia_invalida": [],
    }

    # ---- 1. entrada em carência -------------------------------------
    linha_p = wsp.max_row
    for emp in empresas:
        if emp["numero"] in ja_na_carteira:
            r["ja_na_carteira"].append(emp["numero"])
            continue
        if emp["numero"] in ja_em_pendentes:
            continue

        competencia = emp["competencia"] or referencia
        try:
            validar(competencia)
        except ValueError as exc:
            r["competencia_invalida"].append(f"{emp['numero']}: {exc}")
            continue

        linha_p += 1
        valores = {
            "N° Cliente": emp["numero"], "Nome": emp["nome"], "CNPJ": emp["cnpj"],
            "Regime Tributário": emp["regime"], "Segmento": emp["segmento"],
            "Sugestão Analista": (
                f"{emp['analista']} ({emp['nivel']})" if emp["analista"] and emp["nivel"]
                else emp["analista"]
            ),
            "Origem": ORIGEM_ONBOARDING,
            "Observação": emp["obs"],
            COL_COMPETENCIA: competencia,
            COL_LIBERA: competencia_liberacao(competencia),
        }
        for coluna, col_idx in idx_pend.items():
            if coluna in valores and valores[coluna]:
                wsp.cell(row=linha_p, column=col_idx, value=valores[coluna])
        r["entraram_carencia"].append(f"{emp['numero']} (libera em {valores[COL_LIBERA]})")
        ja_em_pendentes.add(emp["numero"])

    # ---- 2. distribuição de quem venceu a carência ------------------
    # Varre de cima para baixo para a carteira sair na mesma ordem da aba
    # de pendentes; só depois apaga as linhas, de baixo para cima, para os
    # índices não escorregarem durante a remoção.
    a_distribuir: list[tuple[int, dict]] = []

    for linha in range(2, wsp.max_row + 1):
        numero = wsp.cell(row=linha, column=idx_pend["N° Cliente"]).value
        if numero in (None, ""):
            continue
        numero = str(numero).strip()
        entrada = wsp.cell(row=linha, column=idx_pend[COL_COMPETENCIA]).value

        if not entrada:
            # Linha antiga, anterior ao controle de carência: sem competência
            # registrada não dá para saber quando vence — não mexemos nela.
            continue
        entrada = str(entrada).strip()
        try:
            validar(entrada)
        except ValueError as exc:
            r["competencia_invalida"].append(f"{numero}: {exc}")
            continue

        if not liberada(entrada, referencia):
            faltam = competencias_restantes(entrada, referencia)
            verbo = "faltam" if faltam > 1 else "falta"
            r["em_carencia"].append(
                f"{numero} (libera em {competencia_liberacao(entrada)}, "
                f"{verbo} {faltam} competência{'s' if faltam > 1 else ''})"
            )
            continue

        emp = por_numero.get(numero)
        if not emp or not emp["analista"]:
            r["liberadas_sem_responsavel"].append(
                f"{numero} (liberada desde {competencia_liberacao(entrada)})"
            )
            continue

        a_distribuir.append((linha, emp))

    linha_c = ws.max_row
    for _, emp in a_distribuir:
        linha_c += 1
        valores = {
            "N° Cliente": emp["numero"], "Nome": emp["nome"], "CNPJ": emp["cnpj"],
            "Regime Tributário": emp["regime"], "Segmento": emp["segmento"],
            "Analista Responsável": emp["analista"], "Nível": emp["nivel"],
            "Status": STATUS_NOVO,
        }
        for coluna, col_idx in idx_cart.items():
            if coluna in valores:
                ws.cell(row=linha_c, column=col_idx, value=valores[coluna])
        r["distribuidas"].append(emp["numero"])

    for linha, _ in reversed(a_distribuir):
        wsp.delete_rows(linha)

    _estender_filtro(ws, linha_c)
    _estender_filtro(wsp, wsp.max_row)

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    r["saida"] = saida
    r["referencia"] = referencia
    return r


def _bloco(titulo: str, itens: list) -> None:
    if itens:
        print(f"\n{titulo} ({len(itens)}):")
        for i in itens:
            print(f"  - {i}")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    ref = next((a.split("=", 1)[1] for a in sys.argv[1:] if a.startswith("--competencia=")), None)

    if len(args) < 2:
        print(
            "Uso: python -m onboarding.alimentar_carteira "
            "<particularidades.xlsx> <carteira.xlsx> [saida.xlsx] [--competencia=MM/AAAA]"
        )
        raise SystemExit(1)

    part, cart = Path(args[0]), Path(args[1])
    saida = Path(args[2]) if len(args) > 2 else cart.with_name(
        f"{cart.stem}-atualizada-{dt.date.today().isoformat()}.xlsx"
    )

    r = alimentar(part, cart, saida, ref)

    print(f"Carteira gravada em: {r['saida']}")
    print(f"Competência de referência: {r['referencia']} | carência: {MESES_CARENCIA} competências")
    _bloco("Entraram em carência (Pendentes Daniela)", r["entraram_carencia"])
    _bloco("Distribuídas para a carteira", r["distribuidas"])
    _bloco("Ainda em carência", r["em_carencia"])
    _bloco("Carência vencida, mas SEM responsável definido", r["liberadas_sem_responsavel"])
    _bloco("Já estavam na carteira, ignoradas", r["ja_na_carteira"])
    _bloco("Competência inválida — corrija na planilha", r["competencia_invalida"])
    print("\nAbra no Excel para o 'Resumo Equipe' recalcular as contagens.")
