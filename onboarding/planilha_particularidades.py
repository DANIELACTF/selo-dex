"""Gera a planilha de particularidades — o formulário que o Dep. Fiscal
preenche à mão depois da reunião com o Paulo.

Entrada: `data/empresas_pendentes_distribuicao.csv` (saída da parte 1 — a
triagem do e-mail da Thays).
Saída:   `data/particularidades-<data>.xlsx`, com uma linha por empresa,
         já identificada, e as colunas de preenchimento em branco.

As colunas de decisão têm lista suspensa para evitar digitação livre em
campo que depois vira ficha e alimenta a carteira.
"""
from __future__ import annotations

import csv
import datetime as dt
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "1B3A5C"
CINZA_CAB = "EAEDF0"
AMARELO_PREENCHER = "FFF7E0"
BORDA = Border(*[Side(style="thin", color="B8BEC6")] * 4)

# Equipe do Dep. Fiscal — conferido contra a aba "Resumo Equipe" da
# Carteira Tributária Fiscal. Atualize aqui quando a equipe mudar.
ANALISTAS = [
    "Ana Carolina Giordano",
    "Matheus Telles",
    "Thayane Sabia",
    "Dulce Neves",
    "Alexandre Sabino",
    "Monica Oliveira",
    "Wellington",
    "Daniela Carvalho",
]
NIVEIS = ["Sênior", "Júnior", "Auxiliar", "Gestão Fiscal"]
SITUACOES = ["Pendente distribuição", "Distribuído", "Em implantação", "Ativo"]
SEGMENTOS = ["Comércio", "Serviço", "Indústria", "Misto"]
REGIMES = ["Simples Nacional", "Lucro Presumido", "Lucro Real", "MEI", "⚠ A confirmar"]
SIM_PENDENTE = ["recebido", "pendente"]
SENHA_STATUS = ["arquivada", "pendente"]
PROCURACAO = ["obtida", "pendente"]

# (título, largura, chave de pré-preenchimento, lista de validação)
COLUNAS = [
    ("N° Cliente", 10, "numero", None),
    ("Razão social", 42, "empresa", None),
    ("CNPJ", 20, "cnpj", None),
    ("Nome fantasia", 24, None, None),
    ("Município / UF", 26, None, None),
    ("Inscrição Estadual", 20, None, None),
    ("Inscrição Municipal", 20, None, None),
    ("Certificado A1", 14, "certificado", SIM_PENDENTE),
    ("Validade do cert.", 15, None, None),
    ("Senha (cofre)", 14, None, SENHA_STATUS),
    ("Procuração e-CAC", 16, None, PROCURACAO),
    ("Particularidade 1 (Paulo)", 46, None, None),
    ("Particularidade 2 (Paulo)", 46, None, None),
    ("Particularidade 3 (Paulo)", 46, None, None),
    ("Particularidade 4 (Paulo)", 46, None, None),
    ("Responsável (analista)", 22, None, ANALISTAS),
    ("Nível / equipe", 15, None, NIVEIS),
    ("Situação", 20, None, SITUACOES),
    ("Backup / apoio", 20, None, ANALISTAS),
    ("Segmento", 13, None, SEGMENTOS),
    ("Regime confirmado", 18, "regime", REGIMES),
    ("Obs. para a carteira", 40, None, None),
]

# Colunas 1-3 vêm da triagem e não devem ser reescritas à mão.
COLS_IDENTIFICACAO = 3


def _ler_pendentes(csv_path: Path) -> list[dict]:
    if not csv_path.exists():
        raise SystemExit(
            f"Não encontrei {csv_path}. Rode a triagem primeiro:\n"
            f"    python cli.py --input <email.txt>"
        )
    with csv_path.open(encoding="utf-8", newline="") as f:
        return [linha for linha in csv.DictReader(f) if linha.get("numero")]


def _aba_instrucoes(wb: Workbook) -> None:
    ws = wb.create_sheet("Instruções", 0)
    ws.column_dimensions["A"].width = 108
    linhas = [
        ("PLANILHA DE PARTICULARIDADES — Dep. Fiscal Moraex", True),
        ("", False),
        ("Para que serve: registrar o que foi definido na reunião com o Paulo, por empresa,", False),
        ("para virar a Ficha Cadastral definitiva que vai para a pasta do cliente na rede.", False),
        ("", False),
        ("Como preencher", True),
        ("1. As três primeiras colunas (N° Cliente, Razão social, CNPJ) já vêm da triagem —", False),
        ("   não altere: são elas que ligam esta linha à ficha e à carteira.", False),
        ("2. As células em amarelo são as que você preenche. Campo sem informação, deixe", False),
        ("   em branco — a ficha imprime '—' e não inventa dado.", False),
        ("3. As colunas 'Particularidade 1 a 4 (Paulo)' são as definições da reunião. Elas", False),
        ("   saem na ficha em destaque (▶), acima das particularidades do onboarding (•).", False),
        ("   Escreva uma definição por coluna, em uma frase.", False),
        ("4. Colunas com lista suspensa só aceitam os valores da lista — é o que mantém a", False),
        ("   carteira consistente. Precisa de um valor novo? Ajuste a aba 'Listas'.", False),
        ("5. 'Responsável (analista)' e 'Nível / equipe' definem quem assume a empresa;", False),
        ("   enquanto não houver definição, mantenha Situação = 'Pendente distribuição'.", False),
        ("", False),
        ("Depois de preencher", True),
        ("Suba a planilha no Claude e peça as fichas definitivas. Serão geradas as fichas,", False),
        ("a estrutura de pastas da rede, o roteiro de cadastro no G-Click e as linhas para", False),
        ("a Carteira Tributária Fiscal.", False),
    ]
    for i, (texto, negrito) in enumerate(linhas, start=1):
        c = ws.cell(row=i, column=1, value=texto)
        c.font = Font(bold=negrito, size=12 if i == 1 else 10, color=NAVY if negrito else "000000")
        c.alignment = Alignment(vertical="top", wrap_text=True)


def _aba_listas(wb: Workbook) -> None:
    ws = wb.create_sheet("Listas")
    conjuntos = [
        ("Analistas", ANALISTAS), ("Níveis", NIVEIS), ("Situações", SITUACOES),
        ("Segmentos", SEGMENTOS), ("Regimes", REGIMES), ("Certificado", SIM_PENDENTE),
        ("Senha", SENHA_STATUS), ("Procuração", PROCURACAO),
    ]
    for col, (titulo, valores) in enumerate(conjuntos, start=1):
        ws.cell(row=1, column=col, value=titulo).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(titulo) + 4)
        for linha, v in enumerate(valores, start=2):
            ws.cell(row=linha, column=col, value=v)
    ws.sheet_state = "hidden"


def gerar(csv_path: Path, saida: Path) -> Path:
    empresas = _ler_pendentes(csv_path)
    wb = Workbook()
    wb.remove(wb.active)
    _aba_instrucoes(wb)

    ws = wb.create_sheet("Particularidades")

    for i, (titulo, largura, _, _) in enumerate(COLUNAS, start=1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA
        ws.column_dimensions[get_column_letter(i)].width = largura
    ws.row_dimensions[1].height = 32

    for linha, emp in enumerate(empresas, start=2):
        for i, (_, _, chave, _) in enumerate(COLUNAS, start=1):
            valor = None
            if chave == "numero":
                valor = emp["numero"]
            elif chave == "empresa":
                valor = emp["empresa"]
            elif chave == "cnpj":
                valor = emp["cnpj"]
            elif chave == "certificado":
                valor = "recebido" if emp.get("certificado_recebido") == "Sim" else "pendente"
            elif chave == "regime":
                valor = emp.get("regime_informado") or "⚠ A confirmar"

            c = ws.cell(row=linha, column=i, value=valor)
            c.border = BORDA
            c.alignment = Alignment(vertical="top", wrap_text=i > COLS_IDENTIFICACAO)
            c.font = Font(size=9)
            if i <= COLS_IDENTIFICACAO:
                c.fill = PatternFill("solid", fgColor=CINZA_CAB)
            else:
                c.fill = PatternFill("solid", fgColor=AMARELO_PREENCHER)
        ws.row_dimensions[linha].height = 30

    ultima = len(empresas) + 1
    for i, (_, _, _, lista) in enumerate(COLUNAS, start=1):
        if not lista:
            continue
        dv = DataValidation(
            type="list", formula1='"' + ",".join(lista) + '"',
            allow_blank=True, showDropDown=False,
        )
        letra = get_column_letter(i)
        ws.add_data_validation(dv)
        dv.add(f"{letra}2:{letra}{max(ultima, 2)}")

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{max(ultima, 2)}"

    _aba_listas(wb)
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    return saida


if __name__ == "__main__":
    csv_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/empresas_pendentes_distribuicao.csv")
    saida = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(
        f"data/particularidades-{dt.date.today().isoformat()}.xlsx"
    )
    caminho = gerar(csv_path, saida)
    print(f"Planilha gerada: {caminho}")
